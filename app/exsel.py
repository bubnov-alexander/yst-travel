import sqlite3
from openpyxl import Workbook

def export_sql_to_excel(db_path, excel_path):
    # Подключаемся к базе данных
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Выполняем запрос для извлечения всех данных из таблицы catamaran
    cursor.execute("SELECT * FROM catamaran")
    rows = cursor.fetchall()

    # Получаем имена столбцов
    cursor.execute("PRAGMA table_info(catamaran)")
    columns = [column[1] for column in cursor.fetchall()]

    # Создаем новый рабочий файл Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catamaran Data"

    # Записываем имена столбцов в первую строку
    for col_num, column_name in enumerate(columns, 1):
        sheet.cell(row=1, column=col_num, value=column_name)

    # Записываем данные в таблицу
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_data)

    # Сохраняем файл Excel
    workbook.save(excel_path)

    # Закрываем соединение с базой данных
    conn.close()

