def export_sql_to_excel(db_path, excel_path):
    import sqlite3
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
                   SELECT o.id,
                          o.date_arrival,
                          o.time_arrival,
                          o.date_departure,
                          o.time_departure,
                          r.point_a || ' - ' || r.point_b                                                 AS route_name,
                          o.customer_name,
                          o.phone,
                          o.prepayment_status,
                          o.additional_wishes,

                          -- Количество услуг
                          IFNULL(sup.supboards_count, 0)                                                  AS сапборды,
                          IFNULL((SELECT SUM(quantity) FROM catamaran_services WHERE order_id = o.id), 0) AS катамараны,
                          IFNULL((SELECT COUNT(*) FROM transfer_services WHERE order_id = o.id), 0)       AS трансферы,

                          -- Общая стоимость
                          (
                              IFNULL(sup.price, 0) +
                              IFNULL((SELECT SUM(price) FROM catamaran_services WHERE order_id = o.id), 0) +
                              IFNULL((SELECT SUM(price) FROM transfer_services WHERE order_id = o.id), 0)
                              )                                                                           AS общая_стоимость

                   FROM orders o
                            LEFT JOIN routes r ON o.route_id = r.id
                            LEFT JOIN supboard_services sup ON o.id = sup.order_id
                   GROUP BY o.id
                   """)

    rows = cursor.fetchall()

    headers = [
        "ID заказа",
        "Дата приезда",
        "Время приезда",
        "Дата выезда",
        "Время выезда",
        "Маршрут",
        "ФИО клиента",
        "Телефон",
        "Статус предоплаты",
        "Дополнительные пожелания",
        "Сапборды",
        "Катамараны",
        "Трансферы",
        "Общая стоимость"
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заказы"

    for col_num, column_name in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=column_name)

    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_data)

    for col_num, column_title in enumerate(headers, 1):
        max_length = len(str(column_title))
        for row in sheet.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    workbook.save(excel_path)
    conn.close()
